import numpy,pandas,scipy.io,os.path,shutil,zipfile
from openpyxl import load_workbook
from os import listdir, system
from TimeDuration_v02 import TimeDuration as TimeDuration
from datetime import datetime
def OutputsLoop(Ytest,decision,mfilename,start_time,filezip,FileName,SheetName,ResultsFolder,mfilenameAlg,mfilenameAlgVer,NetworkParameters,BSelecAlgName,AIAlgName,LocalFolder,InputFileNameTr,InputFileNameTe):    
    # breakpoint()
    [n_testsamples,n_decisions] = Ytest.shape
    accuracy=round(numpy.sum(decision==Ytest)/(n_testsamples*n_decisions)*100,4)
    accuracystr=str(accuracy)
    o1=['The M-ary accuracy percentage of ', mfilename, ' is ',accuracystr,'%\n'] # Decimal. Correct version
    #o1=['The binary accuracy percentage of BIId_v02 is ',str(round((1-numpy.mean(abs(decision-Ytest)))*100,2)),'%\n']
    print(o1)       # Binary
    
    mat={}
    mat['decision']=decision
    filemat='Decision.mat'
    scipy.io.savemat(filemat,mat)           
    
    timeduration=TimeDuration(start_time)    
    o2=["The simulation duration is ", timeduration]    
    print(o2),print("\n")
    
    filelog='Results.log'
    fo1 = open(filelog, 'w')
    fo1.writelines(o1)
    fo1.writelines(o2)
    fo1.close()
    # shutil.copy(filelog,ResultsFolder)  
    
    
    NetworkParameters['Alg. Name']=BSelecAlgName+'|'+AIAlgName
    NetworkParameters['Acc. Or.']=accuracy    
    NetworkParameters['Total Time AI']=timeduration
    NetworkParameters['Data Input Location']=InputFileNameTr+','+InputFileNameTe
    NetworkParameters['Data Output Location']=LocalFolder+ResultsFolder[1:]+'\\'+filezip
    now = datetime.now().replace(microsecond=0)
    NetworkParameters['Date']=now    
    
    FileNamex = [x for x in listdir() if x.endswith(".xlsx")]    
    while len(FileNamex)>1:
            print('Please, close the',FileName,'file and then press ENTER \n')      
            system('pause')
            FileNamex = [x for x in listdir() if x.endswith(".xlsx")]

    wb = load_workbook(FileName)
    #writer = pandas.ExcelWriter(FileName, engine='openpyxl') 
    #writer.wb = wb

    if not SheetName in wb.sheetnames:
        SheetFirstRow=pandas.DataFrame(columns=NetworkParameters.columns)
        SheetFirstRow.insert(10,'Acc. New','')
        SheetFirstRow.insert(11,'Sum-Rate','')
        SheetFirstRow.insert(12,'Sum-Rate Orig.','')
        SheetFirstRow.insert(13,'Sum-Rate Orig. Mod.','')
        SheetFirstRow.insert(14,'Cases','')
        SheetFirstRow.insert(15,'BCC Err. Perc.','')    
        SheetFirstRow.insert(16,'Total Time Sum-Rate','')    
        #wb = load_workbook(FileName)    
        with pandas.ExcelWriter(FileName, engine='openpyxl') as writer:
            writer.book = wb
            writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)    
            SheetFirstRow.to_excel(writer, sheet_name=SheetName,index=False)

    output = pandas.read_excel(FileName, sheet_name=SheetName)             
    #writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    with pandas.ExcelWriter(FileName, engine='openpyxl') as writer:
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)    
        NetworkParameters[['L','K','M','Network','Tx dBm','Init','Iter','GTS','Alg. Name','Acc. Or.']][0:1].to_excel(writer, sheet_name=SheetName,index=False,header=False,startrow=len(output)+1)
        NetworkParameters[['Total Time AI','Data Input Location','Data Output Location','Date']][0:1].to_excel(writer, sheet_name=SheetName,index=False,header=False,startrow=len(output)+1,startcol=17)   
    writer.save() # Execute Anaconda as Administrator
    writer.close()
    
    zip = zipfile.ZipFile(filezip,'w')
    zip.write(filemat)
    zip.write(filelog)
    zip.write(FileName)
    zip.write(mfilename)
    shutil.make_archive('library', 'zip', 'library')
    zip.write('library.zip')  
    filelog='Results_HyperParam.log'
    if os.path.isfile(filelog):
        zip.write(filelog)    
    zip.close()
    shutil.copy(filezip,ResultsFolder) 
    
    ## Start - There is a small bug here. Needs to be corrected
    # Copy the results to the code_01 main folder as well
    # CodeFolder=LocalFolder[-2:]
    # if CodeFolder!='01':
    #     #LocalFolderx = LocalFolder[:-7]
    #     MainResultsFolder="."+ResultsFolder
    #     shutil.copy(filezip,MainResultsFolder)     
    ## End

    return writer